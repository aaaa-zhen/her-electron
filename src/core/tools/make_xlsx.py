#!/usr/bin/env python3
"""Generate a polished XLSX from JSON input.

Usage: echo '<json>' | python3 make_xlsx.py <output_path>

JSON schema:
{
  "sheets": [
    {
      "name": "Sheet1",
      "headers": ["Col A", "Col B", "Col C"],
      "rows": [
        ["val1", "val2", "val3"],
        ["val4", "val5", "val6"]
      ],
      "column_widths": [20, 15, 15],
      "freeze": "A2",
      "auto_filter": true,
      "formulas": {
        "D2": "=SUM(B2:C2)",
        "D3": "=SUM(B3:C3)"
      },
      "chart": {
        "type": "bar",
        "title": "Chart Title",
        "data_range": "A1:C5",
        "position": "E2"
      }
    }
  ],
  "theme": "blue" | "green" | "dark" | "minimal"
}
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

THEMES = {
    "blue": {
        "header_fill": "1E3A5F",
        "header_font": "FFFFFF",
        "alt_fill": "EDF2F7",
        "accent": "3B82F6",
        "border": "CBD5E1",
    },
    "green": {
        "header_fill": "065F46",
        "header_font": "FFFFFF",
        "alt_fill": "ECFDF5",
        "accent": "10B981",
        "border": "A7F3D0",
    },
    "dark": {
        "header_fill": "1F2937",
        "header_font": "F9FAFB",
        "alt_fill": "F3F4F6",
        "accent": "6EE7B7",
        "border": "D1D5DB",
    },
    "minimal": {
        "header_fill": "F8FAFC",
        "header_font": "1E293B",
        "alt_fill": "FFFFFF",
        "accent": "64748B",
        "border": "E2E8F0",
    },
}


def apply_theme(ws, sheet_data, theme):
    headers = sheet_data.get("headers", [])
    rows = sheet_data.get("rows", [])
    col_widths = sheet_data.get("column_widths", [])
    formulas = sheet_data.get("formulas", {})

    header_fill = PatternFill(start_color=theme["header_fill"], end_color=theme["header_fill"], fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color=theme["header_font"])
    alt_fill = PatternFill(start_color=theme["alt_fill"], end_color=theme["alt_fill"], fill_type="solid")
    body_font = Font(name="Arial", size=11, color="374151")
    thin_border = Border(
        left=Side(style="thin", color=theme["border"]),
        right=Side(style="thin", color=theme["border"]),
        top=Side(style="thin", color=theme["border"]),
        bottom=Side(style="thin", color=theme["border"]),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Try to convert numbers
            if isinstance(value, str):
                try:
                    if "." in value:
                        value = float(value)
                    else:
                        value = int(value)
                except (ValueError, TypeError):
                    pass
            cell.value = value
            cell.font = body_font
            cell.alignment = left_align if isinstance(value, str) else center_align
            cell.border = thin_border
            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Apply formulas
    for cell_ref, formula in formulas.items():
        ws[cell_ref] = formula
        ws[cell_ref].font = Font(name="Arial", size=11, bold=True, color=theme["accent"])
        ws[cell_ref].border = thin_border
        ws[cell_ref].alignment = center_align

    # Column widths
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-width for columns without explicit width
    if not col_widths:
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 8
            for row in rows:
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(str(row[col_idx - 1])))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze panes
    freeze = sheet_data.get("freeze", "A2")
    if freeze:
        ws.freeze_panes = freeze

    # Auto filter
    if sheet_data.get("auto_filter", True) and headers:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Row height
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 24

    # Chart
    chart_data = sheet_data.get("chart")
    if chart_data:
        add_chart(ws, chart_data, headers, rows)


def add_chart(ws, chart_data, headers, rows):
    chart_type = chart_data.get("type", "bar")
    title = chart_data.get("title", "")
    position = chart_data.get("position", f"{get_column_letter(len(headers) + 2)}2")

    num_rows = len(rows) + 1
    num_cols = len(headers)

    if chart_type == "pie":
        chart = PieChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=num_rows, max_col=2)
        cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    elif chart_type == "line":
        chart = LineChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=num_rows, max_col=num_cols)
        cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 10
    else:
        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=1, max_row=num_rows, max_col=num_cols)
        cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 10

    chart.title = title
    chart.width = 18
    chart.height = 12
    ws.add_chart(chart, position)


def generate(data, output_path):
    wb = Workbook()

    theme_name = data.get("theme", "blue")
    theme = THEMES.get(theme_name, THEMES["blue"])

    sheets = data.get("sheets", [])
    if not sheets:
        sheets = [data]  # Treat root as single sheet

    for i, sheet_data in enumerate(sheets):
        if i == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = sheet_data.get("name", f"Sheet{i + 1}")
        apply_theme(ws, sheet_data, theme)

    wb.save(output_path)
    print(f"OK:{output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: echo '<json>' | python3 make_xlsx.py <output>", file=sys.stderr)
        sys.exit(1)
    raw = sys.stdin.read()
    data = json.loads(raw)
    generate(data, sys.argv[1])
